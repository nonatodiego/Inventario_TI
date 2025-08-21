from flask import Blueprint, jsonify, request, send_file
from src.models.user import User, Asset, db
from src.routes.auth import token_required
from sqlalchemy import or_
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import os
import tempfile
from datetime import datetime

export_bp = Blueprint('export', __name__)

@export_bp.route('/export/excel', methods=['GET'])
@token_required
def export_excel(current_user):
    """Exportar dados para Excel"""
    try:
        # Parâmetros de filtro
        search = request.args.get('search', '')
        setor = request.args.get('setor', '')
        gestor = request.args.get('gestor', '')
        
        # Aplicar filtros
        query = User.query
        
        if search:
            query = query.filter(
                or_(
                    User.nome_usuario.ilike(f'%{search}%'),
                    User.matricula.ilike(f'%{search}%'),
                    User.setor.ilike(f'%{search}%')
                )
            )
        
        if setor:
            query = query.filter(User.setor.ilike(f'%{setor}%'))
        
        if gestor:
            query = query.filter(User.nome_gestor.ilike(f'%{gestor}%'))
        
        users = query.all()
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventário de Ativos de TI"
        
        # Cabeçalhos
        headers = [
            'Nome do Usuário', 'Matrícula', 'Setor', 'Nome do Gestor', 'Localização',
            'Desktop/Notebook', 'Segunda Tela', 'Licença Office', 'Celular Corporativo',
            'Headset', 'Mouse sem Fio', 'Teclado sem Fio'
        ]
        
        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Adicionar cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Adicionar dados
        for row, user in enumerate(users, 2):
            assets = user.assets[0] if user.assets else None
            
            data = [
                user.nome_usuario or '',
                user.matricula or '',
                user.setor or '',
                user.nome_gestor or '',
                user.localizacao or '',
                user.desktop_notebook or '',
                'Sim' if user.segunda_tela else 'Não',
                user.licenca_office or '',
                'Sim' if assets and assets.celular_corporativo else 'Não',
                'Sim' if assets and assets.headset else 'Não',
                'Sim' if assets and assets.mouse_sem_fio else 'Não',
                'Sim' if assets and assets.teclado_sem_fio else 'Não'
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar arquivo temporário
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Gerar nome do arquivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"inventario_ativos_ti_{timestamp}.xlsx"
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/export/pdf', methods=['GET'])
@token_required
def export_pdf(current_user):
    """Exportar dados para PDF"""
    try:
        # Parâmetros de filtro
        search = request.args.get('search', '')
        setor = request.args.get('setor', '')
        gestor = request.args.get('gestor', '')
        
        # Aplicar filtros
        query = User.query
        
        if search:
            query = query.filter(
                or_(
                    User.nome_usuario.ilike(f'%{search}%'),
                    User.matricula.ilike(f'%{search}%'),
                    User.setor.ilike(f'%{search}%')
                )
            )
        
        if setor:
            query = query.filter(User.setor.ilike(f'%{setor}%'))
        
        if gestor:
            query = query.filter(User.nome_gestor.ilike(f'%{gestor}%'))
        
        users = query.all()
        
        # Criar arquivo temporário
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        temp_file.close()
        
        # Criar documento PDF
        doc = SimpleDocTemplate(temp_file.name, pagesize=A4)
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Título
        title = Paragraph("Inventário de Ativos de TI", title_style)
        elements.append(title)
        
        # Data de geração
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1  # Center
        )
        date_text = f"Relatório gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
        date_para = Paragraph(date_text, date_style)
        elements.append(date_para)
        elements.append(Spacer(1, 20))
        
        # Preparar dados da tabela
        data = [['Nome', 'Matrícula', 'Setor', 'Gestor', 'Localização', 'Equipamentos']]
        
        for user in users:
            assets = user.assets[0] if user.assets else None
            
            # Listar equipamentos
            equipamentos = []
            if user.desktop_notebook:
                equipamentos.append(user.desktop_notebook)
            if user.segunda_tela:
                equipamentos.append('Segunda Tela')
            if assets:
                if assets.celular_corporativo:
                    equipamentos.append('Celular')
                if assets.headset:
                    equipamentos.append('Headset')
                if assets.mouse_sem_fio:
                    equipamentos.append('Mouse')
                if assets.teclado_sem_fio:
                    equipamentos.append('Teclado')
            if user.licenca_office:
                equipamentos.append(user.licenca_office)
            
            equipamentos_str = ', '.join(equipamentos) if equipamentos else 'Nenhum'
            
            row = [
                user.nome_usuario or '',
                user.matricula or '',
                user.setor or '',
                user.nome_gestor or '',
                user.localizacao or '',
                equipamentos_str
            ]
            data.append(row)
        
        # Criar tabela
        table = Table(data, colWidths=[1.5*inch, 1*inch, 1.2*inch, 1.2*inch, 1*inch, 2*inch])
        
        # Estilo da tabela
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        
        # Estatísticas
        elements.append(Spacer(1, 30))
        stats_title = Paragraph("Estatísticas", styles['Heading2'])
        elements.append(stats_title)
        
        total_users = len(users)
        celular_count = sum(1 for user in users if user.assets and user.assets[0].celular_corporativo)
        headset_count = sum(1 for user in users if user.assets and user.assets[0].headset)
        segunda_tela_count = sum(1 for user in users if user.segunda_tela)
        
        stats_data = [
            ['Métrica', 'Quantidade'],
            ['Total de Usuários', str(total_users)],
            ['Celulares Corporativos', str(celular_count)],
            ['Headsets', str(headset_count)],
            ['Segunda Tela', str(segunda_tela_count)]
        ]
        
        stats_table = Table(stats_data, colWidths=[3*inch, 1*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(stats_table)
        
        # Construir PDF
        doc.build(elements)
        
        # Gerar nome do arquivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"inventario_ativos_ti_{timestamp}.pdf"
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

