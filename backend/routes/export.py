from flask import Blueprint, request, jsonify, send_file
from models import User, Asset, db
from sqlalchemy import or_
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import io
from datetime import datetime

export_bp = Blueprint('export', __name__)

@export_bp.route('/excel', methods=['GET'])
def export_excel():
    """Exportar dados para Excel"""
    try:
        # Parâmetros de filtro
        search = request.args.get('search', '')
        setor = request.args.get('setor', '')
        gestor = request.args.get('gestor', '')

        query = User.query

        # Aplicar filtros
        if search:
            query = query.filter(
                or_(
                    User.nome_usuario.ilike(f'%{search}%'),
                    User.matricula.ilike(f'%{search}%'),
                    User.setor.ilike(f'%{search}%')
                )
            )

        if setor:
            query = query.filter(User.setor == setor)

        if gestor:
            query = query.filter(User.nome_gestor == gestor)

        users = query.all()

        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventário de Ativos"

        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Cabeçalhos
        headers = [
            "Nome", "Matrícula", "Setor", "Gestor", "Localização",
            "Equipamento", "Segunda Tela", "Licença Office",
            "Celular", "Headset", "Mouse", "Teclado"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Dados
        for row, user in enumerate(users, 2):
            asset = user.assets[0] if user.assets else None
            
            ws.cell(row=row, column=1, value=user.nome_usuario)
            ws.cell(row=row, column=2, value=user.matricula)
            ws.cell(row=row, column=3, value=user.setor or "")
            ws.cell(row=row, column=4, value=user.nome_gestor or "")
            ws.cell(row=row, column=5, value=user.localizacao or "")
            ws.cell(row=row, column=6, value=user.desktop_notebook or "")
            ws.cell(row=row, column=7, value="Sim" if user.segunda_tela else "Não")
            ws.cell(row=row, column=8, value=user.licenca_office or "")
            ws.cell(row=row, column=9, value="Sim" if asset and asset.celular_corporativo else "Não")
            ws.cell(row=row, column=10, value="Sim" if asset and asset.headset else "Não")
            ws.cell(row=row, column=11, value="Sim" if asset and asset.mouse_sem_fio else "Não")
            ws.cell(row=row, column=12, value="Sim" if asset and asset.teclado_sem_fio else "Não")

        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Salvar em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"inventario_ativos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'message': 'Erro ao gerar arquivo Excel'}), 500

@export_bp.route('/pdf', methods=['GET'])
def export_pdf():
    """Exportar dados para PDF"""
    try:
        # Parâmetros de filtro
        search = request.args.get('search', '')
        setor = request.args.get('setor', '')
        gestor = request.args.get('gestor', '')

        query = User.query

        # Aplicar filtros
        if search:
            query = query.filter(
                or_(
                    User.nome_usuario.ilike(f'%{search}%'),
                    User.matricula.ilike(f'%{search}%'),
                    User.setor.ilike(f'%{search}%')
                )
            )

        if setor:
            query = query.filter(User.setor == setor)

        if gestor:
            query = query.filter(User.nome_gestor == gestor)

        users = query.all()

        # Criar PDF em memória
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )

        # Conteúdo
        story = []
        
        # Título
        title = Paragraph("Relatório de Inventário de Ativos de TI", title_style)
        story.append(title)
        story.append(Spacer(1, 20))

        # Informações do relatório
        info_style = styles['Normal']
        info = Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", info_style)
        story.append(info)
        story.append(Paragraph(f"Total de registros: {len(users)}", info_style))
        story.append(Spacer(1, 20))

        # Tabela de dados
        data = [['Nome', 'Matrícula', 'Setor', 'Gestor', 'Equipamento', 'Ativos']]
        
        for user in users:
            asset = user.assets[0] if user.assets else None
            ativos = []
            
            if user.segunda_tela:
                ativos.append("2ª Tela")
            if asset:
                if asset.celular_corporativo:
                    ativos.append("Celular")
                if asset.headset:
                    ativos.append("Headset")
                if asset.mouse_sem_fio:
                    ativos.append("Mouse")
                if asset.teclado_sem_fio:
                    ativos.append("Teclado")
            
            data.append([
                user.nome_usuario,
                user.matricula,
                user.setor or "",
                user.nome_gestor or "",
                user.desktop_notebook or "",
                ", ".join(ativos) if ativos else "Nenhum"
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(table)
        doc.build(story)

        buffer.seek(0)
        filename = f"inventario_ativos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'message': 'Erro ao gerar arquivo PDF'}), 500
