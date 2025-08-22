import os
import sys
from openpyxl import load_workbook

# Adicionar o diretório src ao path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from flask import Flask
from src.models.user import db, User, Asset, Role, AuthUser
from werkzeug.security import generate_password_hash

def create_app():
    app = Flask(__name__)
    app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{os.path.join(os.path.dirname(__file__), 'src', 'database', 'app.db')}"
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    db.init_app(app)
    return app

def populate_database():
    app = create_app()
    
    with app.app_context():
        # Criar todas as tabelas
        db.create_all()
        
        # Criar roles padrão
        admin_role = Role.query.filter_by(name='admin').first()
        if not admin_role:
            admin_role = Role(name='admin')
            db.session.add(admin_role)
        
        consulta_role = Role.query.filter_by(name='consulta').first()
        if not consulta_role:
            consulta_role = Role(name='consulta')
            db.session.add(consulta_role)
        
        db.session.commit()
        
        # Criar usuários de autenticação padrão
        admin_user = AuthUser.query.filter_by(username='admin').first()
        if not admin_user:
            admin_user = AuthUser(
                username='admin',
                password_hash=generate_password_hash('admin123'),
                role_id=admin_role.id
            )
            db.session.add(admin_user)
        
        consulta_user = AuthUser.query.filter_by(username='consulta').first()
        if not consulta_user:
            consulta_user = AuthUser(
                username='consulta',
                password_hash=generate_password_hash('consulta123'),
                role_id=consulta_role.id
            )
            db.session.add(consulta_user)
        
        db.session.commit()
        
        # Carregar dados da planilha (sem pandas)
        excel_file = '/home/ubuntu/upload/Planilha_Ativos_TI.xlsx'
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print('Planilha vazia. Nada a importar.')
            return

        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        index_by_header = {str(h).strip(): i for i, h in enumerate(headers)}

        def get_val(row_vals, name, default=None):
            i = index_by_header.get(name, None)
            if i is None or i >= len(row_vals):
                return default
            v = row_vals[i]
            return v if v is not None else default

        def as_bool(v):
            if v is None:
                return False
            s = str(v).strip().lower()
            return s in {'sim', 'yes', 'true', '1', 'y'} or v is True or v == 1

        total = max(0, len(rows) - 1)
        print(f"Carregando {total} registros da planilha...")

        for idx, r in enumerate(rows[1:]):
            nome_usuario = get_val(r, 'Nome do Usuário')
            if not nome_usuario:
                # Pula linhas sem nome
                continue

            # Verificar se o usuário já existe
            existing_user = User.query.filter_by(nome_usuario=nome_usuario).first()
            if existing_user:
                print(f"Usuário {nome_usuario} já existe, pulando...")
                continue

            # Criar usuário
            user = User(
                nome_usuario=nome_usuario,
                matricula=f"MAT{1000 + idx}",  # Gerar matrícula fictícia (compatível com lógica anterior)
                setor=get_val(r, 'Setor'),
                nome_gestor=get_val(r, 'Nome do Gestor'),
                localizacao=get_val(r, 'Localização'),
                desktop_notebook=get_val(r, 'Desktop / Notebook'),
                segunda_tela=as_bool(get_val(r, 'Segunda Tela')),
                licenca_office=get_val(r, 'Licença de Office')
            )

            db.session.add(user)
            db.session.flush()  # Para obter o ID do usuário

            # Criar ativos
            mouse_teclado = as_bool(get_val(r, 'Mouse e teclado sem fio'))
            asset = Asset(
                user_id=user.id,
                celular_corporativo=as_bool(get_val(r, 'Celular Corporativo')),
                headset=as_bool(get_val(r, 'Headset')),
                mouse_sem_fio=mouse_teclado,
                teclado_sem_fio=mouse_teclado  # Assumindo que mouse e teclado são juntos
            )

            db.session.add(asset)

            print(f"Adicionado: {nome_usuario}")
        
        db.session.commit()
        print("Banco de dados populado com sucesso!")
        print(f"Total de usuários: {User.query.count()}")
        print(f"Total de ativos: {Asset.query.count()}")

if __name__ == '__main__':
    populate_database()

